from django.shortcuts import get_object_or_404, render
# Create your views here.
from django.http import HttpResponse
# views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

from users.models import Profile
from .models import CustomForm, FormField, FormSubmission, FormSubmissionData, NightlyFormHistory
from django.contrib.auth.models import User

@login_required
def create_form(request):
    users = User.objects.all()
    if request.method == "POST":
        title = request.POST['title']
        description = request.POST.get('description', '')
        creator_ids = request.POST.getlist('creators')
        submitter_ids = request.POST.getlist('submitters')
        form = CustomForm.objects.create(title=title, description=description, created_by=request.user)
        form.allowed_creators.set(User.objects.filter(id__in=creator_ids))
        form.allowed_submitters.set(User.objects.filter(id__in=submitter_ids))
        return redirect('add_fields', form_id=form.id)
    return render(request, 'create_form.html', {'users': users})





# @login_required
# def create_form(request):
#     if request.method == 'POST':
#         form = DynamicFormCreateForm(request.POST)
#         if form.is_valid():
#             form_instance = form.save(commit=False)
#             form_instance.created_by = request.user
#             form_instance.save()
#             form.save_m2m()
#             return redirect('add_fields', form_id=form_instance.id)
#     else:
#         form = DynamicFormCreateForm()
#     return render(request, 'forms/create_form.html', {'form': form})

@login_required
def close_form(request, form_id):
    form = get_object_or_404(CustomForm, id=form_id, created_by=request.user)
    form.is_closed = True
    form.save()
    return redirect('add_fields', form_id=form.id)









@login_required
def add_fields(request, form_id):
    form = CustomForm.objects.get(id=form_id)
    if request.method == 'POST':
        label = request.POST['label']
        field_type = request.POST['field_type']
        FormField.objects.create(form=form, label=label, field_type=field_type)
    return render(request, 'add_fields.html', {'form': form})


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@login_required
@csrf_exempt
def submit_form(request, form_id):
    form = CustomForm.objects.get(id=form_id)

    if request.user not in form.allowed_submitters.all():
        return JsonResponse({"status": "error", "message": "شما اجازه ارسال این فرم را ندارید."}, status=403)

    if request.method == "POST":
        submission = FormSubmission.objects.create(form=form, submitted_by=request.user)
        for field in form.fields.all():
            value = request.POST.get(str(field.id), '')
            FormSubmissionData.objects.create(submission=submission, field=field, value=value)
        return JsonResponse({"status": "success", "message": "فرم با موفقیت ارسال شد."})

    return render(request, 'submit_form.html', {'form': form})



# views.py
from django.shortcuts import render, get_object_or_404
from .models import CustomForm, FormSubmission, FormField, FormSubmissionData

def form_results(request):
    forms = CustomForm.objects.all()
    selected_form = None
    submissions = None
    selected_submission = None
    answers = None

    form_id = request.GET.get('form')
    submission_id = request.GET.get('submission')

    if form_id:
        selected_form = get_object_or_404(CustomForm, id=form_id)
        submissions = FormSubmission.objects.filter(form=selected_form)

    if submission_id:
        selected_submission = get_object_or_404(FormSubmission, id=submission_id)
        answers = FormSubmissionData.objects.filter(submission=selected_submission)

    context = {
        'forms': forms,
        'selected_form': selected_form,
        'submissions': submissions,
        'selected_submission': selected_submission,
        'answers': answers,
    }
    return render(request, 'form_results.html', context)





# # views.py
# from django.contrib.auth.decorators import login_required
# from django.shortcuts import render

# @login_required
# def available_forms(request):
#     forms = CustomForm.objects.all()

#     return render(request, 'available_forms.html', {'forms': forms})



# views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import CustomForm

@login_required
def available_forms(request):
    all_forms = CustomForm.objects.all()

    # Add `can_submit` flag to each form for this user
    forms_with_permission = []
    for form in all_forms:
        can_submit = request.user in form.allowed_submitters.all()
        forms_with_permission.append({
            'form': form,
            'can_submit': can_submit,
        })

    return render(request, 'available_forms.html', {'forms': forms_with_permission})






from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import NightlySalesForm, save_with_persian_labels
from .models import NightlyFormModel
from decimal import Decimal  # اضافه کردن این خط



from django.utils import timezone



# تابع کمکی برای تبدیل Decimal به float
def convert_decimals_to_floats(data):
    """تبدیل تمام مقادیر Decimal در دیکشنری به float"""
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, Decimal):
                data[key] = float(value)
            elif isinstance(value, dict):
                convert_decimals_to_floats(value)
    return data

@login_required
def nightly_sales_view(request):
    if request.method == 'POST':
        form = NightlySalesForm(request.POST)
        if form.is_valid():
            # تبدیل Decimal به float قبل از ذخیره
            additional_form_dict = get_data_from_form(request=request)
            merged_dict = {**form.cleaned_data,**additional_form_dict}
            cleaned_data = convert_decimals_to_floats(merged_dict)
            cleaned_data = save_with_persian_labels(cleaned_data)
            NightlyFormModel.objects.create(
                user=request.user,
                data=cleaned_data
            )
            return redirect('success_page')
    else:
        form = NightlySalesForm()


    context = {
        'form': form,
        'names_list': json.dumps(['milad','ali']),
        # ... سایر داده‌ها
    }

        
    return render(request, 'nightly_form.html',context)



def get_data_from_form(request):
    form_data = request.POST
        
    # داده‌های فرم‌های پویا (فرم 1)
    additional_form_dict = {}
    for i in range(1, 10):  # حداکثر 10 فرم
        name = request.POST.get(f'name_{i}')
        value1 = request.POST.get(f'value1_{i}')
        value2 = request.POST.get(f'value2_{i}')
        value3 = request.POST.get(f'value3_{i}')
        value4 = request.POST.get(f'value4_{i}')
        value5 = request.POST.get(f'value5_{i}')
        value6 = request.POST.get(f'value6_{i}')
        value7 = request.POST.get(f'value7_{i}')
        
        if name or value1 or value2 or value3 or value4 or value5 or value6 or value7:
            additional_form_dict.update({f'نام پیک_{i}': name})
            additional_form_dict.update({f'اسنپ_{i}': value1})
            additional_form_dict.update({f'تلفنی_{i}': value2})
            additional_form_dict.update({f'جمع کارکرد_{i}': value3})
            additional_form_dict.update({f'کمیسیون_{i}': value4})
            additional_form_dict.update({f'غذا_{i}': value5})
            additional_form_dict.update({f'انعام_{i}': value6})
            additional_form_dict.update({f'خالص پرداخت_{i}': value7})
    

    for i in range(1, 10):
        f2_value1 = request.POST.get(f'f2_value1_{i}')
        f2_value2 = request.POST.get(f'f2_value2_{i}')

        
        if f2_value1 or f2_value2 :
            additional_form_dict.update({f'شرح_{i}': f2_value1})
            additional_form_dict.update({f'مبلغ - ریال_{i}': f2_value2})

    


    return additional_form_dict


@login_required
def form_detail(request, form_id):
    form_instance = get_object_or_404(NightlyFormModel, id=form_id, user=request.user)
    now = timezone.now()
    time_diff = now - form_instance.created_at
    is_editable = time_diff.total_seconds() < 7200 or request.user.is_staff

    if request.method == 'POST' and is_editable:
        form = NightlySalesForm(request.POST)
        if form.is_valid():
            # تبدیل Decimal به float قبل از ذخیره
            cleaned_data = convert_decimals_to_floats(form.cleaned_data)
            form_instance.data = cleaned_data
            form_instance.save()
            
            NightlyFormHistory.objects.create(
                form=form_instance,
                user=request.user,
                old_data=form_instance.data,
                new_data=cleaned_data
            )
            return redirect('form_detail', form_id=form_id)
    else:
        # تبدیل Decimal به float برای نمایش در فرم
        initial_data = convert_decimals_to_floats(form_instance.data)
        form = NightlySalesForm(initial=initial_data)

    return render(request, 'form_detail.html', {
        'form': form,
        'form_instance': form_instance,
        'is_editable': is_editable
    })







from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.views.generic import ListView
from django.urls import reverse_lazy
from .models import NightlyFormModel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import json

class NightlyFormListView(ListView):
    model = NightlyFormModel
    template_name = 'nightly_forms_list.html'
    context_object_name = 'forms'
    ordering = ['-created_at']
    paginate_by = 30

    def get_queryset(self):
        return NightlyFormModel.objects.filter(user=self.request.user).order_by('-created_at')[:30]

def download_excel(request, form_id):
    form = get_object_or_404(NightlyFormModel, id=form_id, user=request.user)
    
    # ایجاد کتابکار Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "فرم شبانه"
    
    # اضافه کردن سرستون‌ها
    headers = list(form.data.keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    # اضافه کردن داده‌ها
    for col_num, value in enumerate(form.data.values(), 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # تنظیمات خروجی
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # ایجاد پاسخ دانلود
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=form_{form.id}_{form.date}.xlsx'
    return response





def peyk_create(request):
    return render(request, 'form_peyk.html')



# views.py
from django.http import JsonResponse

def get_people(request):
    people = Profile.objects.all().values('id', 'name')
    
    return JsonResponse(list(people), safe=False)
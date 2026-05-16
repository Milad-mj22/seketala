import jdatetime
from datetime import datetime
import re
def jalali_date_time_to_gregorian(jalali_date, time_str):
    """
    jalali_date: '1404-10-12'
    time_str: '14:32:10'
    """
    jy, jm, jd = map(int, jalali_date.split("-"))
    hour, minute = map(int, time_str.split(":"))

    j_date = jdatetime.date(jy, jm, jd)
    g_date = j_date.togregorian()

    return datetime(
        g_date.year,
        g_date.month,
        g_date.day,
        hour,
        minute,
       
    )



import re

def extract_payment_methods(nahveh_text):
    """
    استخراج روش پرداخت از رشته nahveh
    - اگر [] وجود داشته باشد، محتوا داخل براکت برگردد
    - اگر [] نبود، اگر شامل 'نقدي' بود فقط 'نقدي' برگردد
    - در غیر اینصورت کل متن را برگرداند
    """
    if not nahveh_text:
        return set()

    # 1️⃣ اگر براکت وجود داشته باشد
    if '[' in nahveh_text and ']' in nahveh_text:
        matches = re.findall(r'\[\s*(.*?)\s*\]', nahveh_text)
        return set(matches) if matches else set()
    
    # 2️⃣ اگر شامل "نقدي" بود
    elif 'نقدي' in nahveh_text:
        return {'نقدي'}
    
    # 3️⃣ در غیر اینصورت کل متن
    else:
        return {nahveh_text.strip()}



from datetime import datetime, time, timedelta
from django.utils import timezone

from datetime import datetime, time, timedelta
from django.utils import timezone


def get_date_range(selected_date):
    # Start at 03:00 of selected date
    start = datetime.combine(selected_date, time(7, 0))
    # End at 03:00 of next day
    end = start + timedelta(days=10)

    return start, end


def get_date_range_night_form(selected_date):
    # Start at 03:00 of selected date
    start = datetime.combine(selected_date, time(3, 0))

    # End at 03:00 of next day
    end = start + timedelta(days=-1)

    return end, start


from user_management.utils import check_server
SERVER = check_server()
from openpyxl import load_workbook


def read_excel(excel_name:str):

    if SERVER:
        template_path = f"/home/seketal1/Seketala_Kitchen_Flow/cache/{excel_name}"  # must be .xlsx
    else:
        template_path = f'cache\{excel_name}'


        # بارگذاری فایل و خواندن داده‌ها
    wb = load_workbook(template_path)
    ws =  wb[wb.sheetnames[0]] # یا wb["نام_سیت"] برای انتخاب سیت خاص

    # خواندن تمام داده‌های سیت
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    return data



def get_persian_date_string(date_obj):
    """تبدیل تاریخ میلادی به رشته شمسی فارسی"""
    import jdatetime
    
    persian_date = jdatetime.date.fromgregorian(date=date_obj)
    
    days_name = [ 'شنبه', 'یکشنبه','دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    months_name = ['', 'فروردین', 'اردیبهشت', 'خرداد', 'تیر', 'مرداد', 'شهریور', 
                   'مهر', 'آبان', 'آذر', 'دی', 'بهمن', 'اسفند']
    
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    
    def to_persian(num):
        return ''.join(persian_digits[int(d)] for d in str(num))
    
    return f"{days_name[persian_date.weekday()]} {to_persian(persian_date.day)} {months_name[persian_date.month]} {to_persian(persian_date.year)}"